import re
import pandas as pd
import sys
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

# ── Color Themes ──────────────────────────────────────
THEMES = {
    "blue": {
        "name": "Professional Blue",
        "header_fill": "4F81BD",
        "header_font": "FFFFFF",
        "title_fill": "1F497D",
        "title_font": "FFFFFF",
        "accent": "DCE6F1",
    },
    "green": {
        "name": "Forest Green",
        "header_fill": "548235",
        "header_font": "FFFFFF",
        "title_fill": "375623",
        "title_font": "FFFFFF",
        "accent": "E2EFDA",
    },
    "sunset": {
        "name": "Warm Sunset",
        "header_fill": "C55A11",
        "header_font": "FFFFFF",
        "title_fill": "843C0C",
        "title_font": "FFFFFF",
        "accent": "FCE4D6",
    },
    "dark": {
        "name": "Dark Mode",
        "header_fill": "333333",
        "header_font": "E0E0E0",
        "title_fill": "1A1A1A",
        "title_font": "FFFFFF",
        "accent": "444444",
    },
    "purple": {
        "name": "Soft Purple",
        "header_fill": "7030A0",
        "header_font": "FFFFFF",
        "title_fill": "4A1A6B",
        "title_font": "FFFFFF",
        "accent": "E4CCFF",
    },
}

def parse_transcript(file_path):
    with open(file_path, "r", encoding="utf-8") as f:
        lines = f.readlines()

    data = []
    
    # Regex to match the speaker turn format: "S01: text" or "S00: text"
    speaker_pattern = re.compile(r"^(S\d{2}):\s*(.*)")
    
    current_speaker = None
    current_text = ""
    
    # Skip the header lines
    start_parsing = False
    for line in lines:
        line = line.strip()
        if not line:
            continue
            
        if line == "(..)" or line == "(.)":
            start_parsing = True
            
        if not start_parsing and not speaker_pattern.match(line):
            continue
            
        start_parsing = True
        
        match = speaker_pattern.match(line)
        if match:
            # Save previous turn if it exists
            if current_speaker is not None:
                data.append({
                    "Speaker": current_speaker,
                    "Text": current_text.strip()
                })
            
            # Start new turn
            current_speaker = match.group(1)
            current_text = match.group(2)
        else:
            # Append to current turn
            if current_speaker is not None:
                current_text += " " + line
                
    # Append the last turn
    if current_speaker is not None:
        data.append({
            "Speaker": current_speaker,
            "Text": current_text.strip()
        })
        
    # Extract features for the excel file
    for row in data:
        text = row["Text"]
        row["Word Count"] = len(text.split())
        
        # Count overlaps (indicated by //text//)
        overlaps = len(re.findall(r"//(.*?)//", text))
        row["Contains Overlap?"] = "Yes" if overlaps > 0 else "No"
        
        # Count pauses
        short_pauses = text.count("(.)")
        long_pauses = text.count("(..)") + text.count("(...)")
        row["Short Pauses"] = short_pauses
        row["Long Pauses"] = long_pauses
        
        # Simple disfluency count
        disfluency_pattern = re.compile(r"\b(uh|um|hmm|yeah|like)\b", re.IGNORECASE)
        row["Disfluencies (uh/um/yeah...)"] = len(disfluency_pattern.findall(text))

    return data

def apply_excel_styling(output_file, df, theme_name="blue"):
    theme = THEMES.get(theme_name, THEMES["blue"])
    
    wb = load_workbook(output_file)
    ws = wb.active
    ws.title = "Transcript Data"
    
    # --- 1. Define Visual Styles (from theme) ---
    header_fill = PatternFill(start_color=theme["header_fill"], end_color=theme["header_fill"], fill_type="solid")
    header_font = Font(color=theme["header_font"], bold=True)
    accent_fill = PatternFill(start_color=theme["accent"], end_color=theme["accent"], fill_type="solid")
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    top_alignment = Alignment(vertical="top")
    
    # --- 2. Apply Header Styling and Freeze Top Row ---
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    ws.freeze_panes = "A2"
    
    # --- 3. Set Column Widths ---
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 13
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 25

    # --- 4. Apply Row Formatting with Zebra Striping ---
    for row_idx in range(2, ws.max_row + 1):
        speaker_name = ws.cell(row=row_idx, column=1).value
        ws.cell(row=row_idx, column=2).alignment = wrap_alignment
        for col_idx in [1, 3, 4, 5, 6, 7]:
            ws.cell(row=row_idx, column=col_idx).alignment = top_alignment

        # Zebra striping using theme accent color
        if row_idx % 2 == 0:
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = accent_fill
                
    # --- 5. CREATE THE RESULTS SHEET ---
    ws_results = wb.create_sheet(title="Results Summary", index=0)
    
    # Calculate stats for all speakers
    speakers = list(df["Speaker"].unique())
    stats = {}
    
    # Check if lexical metrics should be computed
    compute_lexical = "Flesch-Kincaid" in df.columns
    
    for s in speakers:
        speaker_df = df[df["Speaker"] == s]
        stats[s] = {
            "Total Turns": len(speaker_df),
            "Total Words": int(speaker_df["Word Count"].sum()),
            "Avg Words per Turn": round(speaker_df["Word Count"].mean(), 1) if len(speaker_df) > 0 else 0,
            "Total Overlaps": len(speaker_df[speaker_df["Contains Overlap?"] == "Yes"]),
            "Total Short Pauses (.)": int(speaker_df["Short Pauses"].sum()),
            "Total Long Pauses (..)": int(speaker_df["Long Pauses"].sum()),
            "Total Disfluencies (uh/yeah)": int(speaker_df["Disfluencies (uh/um/yeah...)"].sum()),
        }
        
        if compute_lexical:
            stats[s]["TTR (Type-Token Ratio)"] = round(speaker_df["TTR"].mean(), 3) if len(speaker_df) > 0 else 0
            stats[s]["MTLD (Lexical Diversity)"] = round(speaker_df["MTLD"].mean(), 1) if len(speaker_df) > 0 else 0
            stats[s]["Flesch-Kincaid Readability"] = round(speaker_df["Flesch-Kincaid"].mean(), 1) if len(speaker_df) > 0 else 0

    # Title styling (from theme)
    title_font = Font(size=16, bold=True, color=theme["title_font"])
    title_fill = PatternFill(start_color=theme["title_fill"], end_color=theme["title_fill"], fill_type="solid")
    
    # Write Title
    ws_results.merge_cells('B2:E2')
    title_cell = ws_results.cell(row=2, column=2, value="SPEAKER COMPARISON FINDINGS")
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write Headers Dynamically
    sub_header_font = Font(bold=True)
    col_idx = 3
    for s in speakers:
        head_cell = ws_results.cell(row=4, column=col_idx, value=str(s).upper())
        head_cell.font = sub_header_font
        head_cell.alignment = Alignment(horizontal="center")
        col_idx += 1

    # Metrics Layout
    metrics = [
        ("Total Speaking Turns", "Total Turns"),
        ("Total Words Spoken", "Total Words"),
        ("Average Words per Turn", "Avg Words per Turn"),
        ("Overlapping Statements", "Total Overlaps"),
        ("Short Pauses (.)", "Total Short Pauses (.)"),
        ("Long Pauses (..)", "Total Long Pauses (..)"),
        ("Disfluencies (uh, um, yeah)", "Total Disfluencies (uh/yeah)")
    ]
    
    if compute_lexical:
        metrics.extend([
            ("TTR (Type-Token Ratio, 0-1)", "TTR (Type-Token Ratio)"),
            ("MTLD (Lexical Diversity, higher=richer)", "MTLD (Lexical Diversity)"),
            ("Flesch-Kincaid Grade Level", "Flesch-Kincaid Readability")
        ])

    current_row = 6
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for display_name, dict_key in metrics:
        # Label
        label_cell = ws_results.cell(row=current_row, column=2, value=display_name)
        label_cell.font = Font(bold=True)
        label_cell.alignment = Alignment(horizontal="right")
        
        col_idx = 3
        for s in speakers:
            stat_cell = ws_results.cell(row=current_row, column=col_idx, value=stats[s][dict_key])
            stat_cell.border = thin_border
            stat_cell.alignment = Alignment(horizontal="center")
            col_idx += 1
        
        current_row += 2 # Double space for presentation

    # Set column widths for results
    ws_results.column_dimensions['A'].width = 5
    ws_results.column_dimensions['B'].width = 30
    ws_results.column_dimensions['C'].width = 20
    ws_results.column_dimensions['D'].width = 20
    ws_results.column_dimensions['E'].width = 5

    wb.save(output_file)

def add_question_column(df):
    """Add a column flagging turns that contain questions."""
    df["Contains Question?"] = df["Text"].apply(lambda t: "Yes" if "?" in t else "No")
    return df


def create_word_frequency_sheet(wb, df, theme_name="blue"):
    """Create a sheet with top words per speaker."""
    from collections import Counter
    import string

    theme = THEMES.get(theme_name, THEMES["blue"])
    ws = wb.create_sheet(title="Word Frequency")

    header_fill = PatternFill(start_color=theme["header_fill"], end_color=theme["header_fill"], fill_type="solid")
    header_font = Font(color=theme["header_font"], bold=True)

    speakers = list(df["Speaker"].unique())
    stop_words = {"the","a","an","and","or","but","in","on","at","to","for","of","is","it","i","you","he","she",
                  "we","they","that","this","was","were","are","be","been","being","have","has","had","do","does",
                  "did","will","would","could","should","may","might","can","shall","so","if","then","than",
                  "not","no","my","your","our","their","its","me","him","her","us","them","what","which","who",
                  "where","when","how","just","also","very","with","from","about","as","all","up","out"}

    col = 1
    for speaker in speakers:
        ws.cell(row=1, column=col, value=f"{speaker} — Word").fill = header_fill
        ws.cell(row=1, column=col, value=f"{speaker} — Word").font = header_font
        ws.cell(row=1, column=col+1, value="Count").fill = header_fill
        ws.cell(row=1, column=col+1, value="Count").font = header_font

        texts = " ".join(df[df["Speaker"] == speaker]["Text"].tolist()).lower()
        words = [w.strip(string.punctuation) for w in texts.split() if w.strip(string.punctuation)]
        words = [w for w in words if w and w not in stop_words and len(w) > 1
                 and w not in {"uh","um","hmm","yeah","like","(.)","(..)","(...)"}]
        top = Counter(words).most_common(30)

        for i, (word, count) in enumerate(top):
            ws.cell(row=i+2, column=col, value=word)
            ws.cell(row=i+2, column=col+1, value=count)

        ws.column_dimensions[get_column_letter(col)].width = 18
        ws.column_dimensions[get_column_letter(col+1)].width = 10
        col += 3  # gap column


def export_latex_tables(df, output_file):
    """Export summary statistics as a LaTeX .tex file ready to paste into a paper."""
    speakers = list(df["Speaker"].unique())
    stats = {}
    for s in speakers:
        sdf = df[df["Speaker"] == s]
        stats[s] = {
            "Total Turns": len(sdf),
            "Total Words": int(sdf["Word Count"].sum()),
            "Avg Words/Turn": round(sdf["Word Count"].mean(), 1) if len(sdf) > 0 else 0,
            "Overlaps": len(sdf[sdf["Contains Overlap?"] == "Yes"]),
            "Short Pauses": int(sdf["Short Pauses"].sum()),
            "Long Pauses": int(sdf["Long Pauses"].sum()),
            "Disfluencies": int(sdf["Disfluencies (uh/um/yeah...)"].sum()),
        }
        if "Contains Question?" in df.columns:
            stats[s]["Questions"] = len(sdf[sdf["Contains Question?"] == "Yes"])

    tex_path = output_file.replace(".xlsx", "_tables.tex")

    with open(tex_path, "w", encoding="utf-8") as f:
        f.write("% Auto-generated by Audio Analysis Pipeline\n")
        f.write("% Paste this into your LaTeX document\n")
        f.write("% Requires: \\usepackage{booktabs}\n\n")

        # Speaker comparison table
        cols = "l" + "r" * len(speakers)
        f.write("\\begin{table}[h]\n")
        f.write("\\centering\n")
        f.write("\\caption{Speaker Comparison Summary}\n")
        f.write("\\label{tab:speaker-comparison}\n")
        f.write(f"\\begin{{tabular}}{{{cols}}}\n")
        f.write("\\toprule\n")

        # Header row
        header = "Metric & " + " & ".join(str(s) for s in speakers) + " \\\\\n"
        f.write(header)
        f.write("\\midrule\n")

        # Data rows
        metrics = [
            ("Total Turns", "Total Turns"),
            ("Total Words", "Total Words"),
            ("Avg Words/Turn", "Avg Words/Turn"),
            ("Overlaps", "Overlaps"),
            ("Short Pauses", "Short Pauses"),
            ("Long Pauses", "Long Pauses"),
            ("Disfluencies", "Disfluencies"),
        ]
        if "Contains Question?" in df.columns:
            metrics.append(("Questions", "Questions"))
        if "Flesch-Kincaid" in df.columns:
            metrics.extend([
                ("TTR", "TTR"),
                ("MTLD", "MTLD"),
                ("F-K Grade", "Flesch-Kincaid")
            ])
            for s in speakers:
                speaker_df = df[df["Speaker"] == s]
                stats[s]["TTR"] = round(speaker_df["TTR"].mean(), 3) if len(speaker_df) > 0 else 0
                stats[s]["MTLD"] = round(speaker_df["MTLD"].mean(), 1) if len(speaker_df) > 0 else 0
                stats[s]["Flesch-Kincaid"] = round(speaker_df["Flesch-Kincaid"].mean(), 1) if len(speaker_df) > 0 else 0

        for display, key in metrics:
            vals = " & ".join(str(stats[s].get(key, 0)) for s in speakers)
            f.write(f"{display} & {vals} \\\\\n")

        f.write("\\bottomrule\n")
        f.write("\\end{tabular}\n")
        f.write("\\end{table}\n")

        # Figure inclusion stubs
        f.write("\n\n% ── Figure stubs (adjust paths as needed) ──\n")
        basename = output_file.replace(".xlsx", "").split("/")[-1].split("\\\\")[-1]
        for fig_file, caption in [
            ("conversational_timeline", "Conversational rhythm: word count per speaking turn"),
            ("friction_heatmap", "Conversational friction: disfluencies and pauses over time"),
            ("speaker_balance", "Speaker balance: share of words and turns"),
            ("turn_taking_matrix", "Turn-taking adjacency matrix: who responds to whom"),
            ("turn_distribution", "Turn length distribution by speaker"),
        ]:
            f.write(f"\n% \\begin{{figure}}[h]\n")
            f.write(f"%   \\centering\n")
            f.write(f"%   \\includegraphics[width=\\textwidth]{{figures/{fig_file}.pdf}}\n")
            f.write(f"%   \\caption{{{caption}}}\n")
            f.write(f"%   \\label{{fig:{fig_file}}}\n")
            f.write(f"% \\end{{figure}}\n")

    return tex_path


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Parse transcript to styled Excel")
    parser.add_argument("input_file", help="Input transcript .txt file")
    parser.add_argument("output_file", help="Output .xlsx file")
    parser.add_argument("--theme", default="blue", choices=list(THEMES.keys()),
                        help="Color theme")
    parser.add_argument("--summary", action="store_true", default=True,
                        help="Include Results Summary sheet (default: on)")
    parser.add_argument("--no-summary", action="store_false", dest="summary",
                        help="Skip Results Summary sheet")
    parser.add_argument("--word-freq", action="store_true", default=False,
                        help="Include Word Frequency sheet")
    parser.add_argument("--lexical", action="store_true", default=False,
                        help="Compute lexical diversity metrics (TTR, MTLD, Readability)")
    parser.add_argument("--questions", action="store_true", default=False,
                        help="Add question detection column")
    parser.add_argument("--csv", action="store_true", default=False,
                        help="Also export a .csv alongside the .xlsx")
    parser.add_argument("--latex", action="store_true", default=False,
                        help="Export summary as LaTeX tables (.tex)")

    args = parser.parse_args()

    if args.theme not in THEMES:
        print(f"Unknown theme '{args.theme}'. Available: {', '.join(THEMES.keys())}")
        return

    print(f"Parsing {args.input_file}...")
    print(f"Using theme: {THEMES[args.theme]['name']}")
    data = parse_transcript(args.input_file)

    df = pd.DataFrame(data)

    # Optional: Lexical metrics
    if args.lexical:
        print("  + Computing lexical diversity (TTR, MTLD, Flesch-Kincaid)...")
        from lexicalrichness import LexicalRichness
        from textstat import textstat
        
        ttrs, mtlds, fks = [], [], []
        for text in df["Text"]:
            # Clean text for analysis (remove annotations)
            clean_text = text.replace("(.)", "").replace("(..)", "").replace("(...)", "").replace("//", "")
            
            # Readability
            try:
                fk = textstat.flesch_kincaid_grade(clean_text)
            except Exception:
                fk = 0
            fks.append(fk)
            
            # Lexical Richness
            try:
                # LexicalRichness fails on very short texts (< 10 words usually)
                lex = LexicalRichness(clean_text)
                ttrs.append(lex.ttr)
                mtlds.append(lex.mtld(threshold=0.72))
            except Exception:
                ttrs.append(0)
                mtlds.append(0)
                
        df["TTR"] = ttrs
        df["MTLD"] = mtlds
        df["Flesch-Kincaid"] = fks

    # Optional: question detection column
    if args.questions:
        df = add_question_column(df)
        columns = ["Speaker", "Text", "Word Count", "Contains Overlap?",
                   "Short Pauses", "Long Pauses", "Disfluencies (uh/um/yeah...)", "Contains Question?"]
        print("  + Question detection column")
    else:
        columns = ["Speaker", "Text", "Word Count", "Contains Overlap?",
                   "Short Pauses", "Long Pauses", "Disfluencies (uh/um/yeah...)"]

    if args.lexical:
        columns.extend(["TTR", "MTLD", "Flesch-Kincaid"])

    df = df[columns]

    print(f"Exporting raw data to {args.output_file}...")
    df.to_excel(args.output_file, index=False, engine="openpyxl")

    print("Applying visual magic and creating Results sheet...")
    apply_excel_styling(args.output_file, df, theme_name=args.theme)

    # Optional: word frequency sheet
    if args.word_freq:
        from openpyxl import load_workbook as lwb
        wb = lwb(args.output_file)
        create_word_frequency_sheet(wb, df, args.theme)
        wb.save(args.output_file)
        print("  + Word Frequency sheet added")

    # Optional: CSV export
    if args.csv:
        csv_path = args.output_file.replace(".xlsx", ".csv")
        df.to_csv(csv_path, index=False, encoding="utf-8")
        print(f"  + CSV exported: {csv_path}")

    # Optional: LaTeX export
    if args.latex:
        tex_path = export_latex_tables(df, args.output_file)
        print(f"  + LaTeX tables exported: {tex_path}")

    print("Done!")

if __name__ == "__main__":
    main()
